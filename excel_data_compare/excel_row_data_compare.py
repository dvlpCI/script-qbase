'''
Author: dvlproad
Date: 2023-10-24 18:04:20
LastEditors: dvlproad dvlproad@163.com
LastEditTime: 2023-10-24 22:27:20
Description: excel 行数据的比较: 依次对行比较，当当前行与其上一行相比，若两个类一致，且tti数据相差超过
'''

# 定义颜色常量
NC='\033[0m' # No Color
RED='\033[31m'
GREEN='\033[32m'
YELLOW='\033[33m'
BLUE='\033[34m'
PURPLE='\033[0;35m'
CYAN='\033[0;36m'

import os
# 获取具名参数的值
import argparse
from plistlib import InvalidFileException
parser = argparse.ArgumentParser()  # 创建参数解析器
parser.add_argument("-filePath", "--filePath", help="The value for argument 'filePath'")
parser.add_argument("-startRowNo", "--startRowNo", help="The value for argument 'startRowNo'")
parser.add_argument("-idColumnNo", "--idColumnNo", help="The value for argument 'idColumnNo'")
parser.add_argument("-valueColumnNo", "--valueColumnNo", help="The value for argument 'valueColumnNo'")
parser.add_argument("-valueDiffColumnNo", "--valueDiffColumnNo", help="The value for argument 'valueDiffColumnNo'")
parser.add_argument("-successMS", "--successMS", help="The value for argument 'successMS'")
parser.add_argument("-failureMS", "--failureMS", help="The value for argument 'failureMS'")
parser.add_argument("-resultSaveToFilePath", "--resultSaveToFilePath", help="The value for argument 'resultSaveToFilePath'")
args = parser.parse_args()  # 解析命令行参数
filePath = args.filePath
successMS = args.successMS
failureMS = args.failureMS
resultSaveToFilePath = args.resultSaveToFilePath
if filePath is None:
    print(f"{RED}您要比较的excel文件 -filePath 不能为空，请检查！{NC}")
    exit(1)
# elif not os.path.exists(filePath):
#     print(f"{RED}您要比较的excel文件 -filePath 参数{BLUE} {filePath} {RED}不存在，请检查！{NC}")
#     exit(1)

if args.startRowNo is None:
    print(f"{RED}从哪一行开始比较 -startRowNo 不能为空，请检查！{NC}")
    exit(1)
else:
    # startRowNo = int(args.startRowNo)
    try:
        startRowNo = int(args.startRowNo)
    except ValueError as e:
        print(f"{RED}startRowNo 类型转换错误:{e} {NC}")
        exit(1)
    except Exception as e:
        print(f"{RED}startRowNo 其他错误:{e} {NC}")
        exit(1)

if args.idColumnNo is None:
    print(f"{RED}判断时候使用哪个列的值作为id参数 -idColumnNo 不能为空，请检查！{NC}")
    exit(1)
else:
    idColumnNo = int(args.idColumnNo)

if args.valueColumnNo is None:
    print(f"{RED}判断时候使用哪个列的值作为value参数 -valueColumnNo 不能为空，请检查！{NC}")
    exit(1)
else:
    valueColumnNo = int(args.valueColumnNo)

if args.valueDiffColumnNo is None:
    print(f"{RED}判断时候value比较的结果放在哪一列 -valueDiffColumnNo 不能为空，请检查！{NC}")
    exit(1)
else:
    valueDiffColumnNo = int(args.valueDiffColumnNo)
    
if successMS is None:
    print(f"{RED}新版本加载时长降低多少则是优化成功的参数(正数) -successMS 不能为空，请检查！{NC}")
    exit(1)
if failureMS is None:
    print(f"{RED}新版本加载时长增加多少则是优化失败的参数(负数) -failureMS 不能为空，请检查！{NC}")
    exit(1)
if resultSaveToFilePath is None:
    print(f"{RED}结果要保存的文件 -resultSaveToFilePath 不能为空，请检查！{NC}")
    exit(1)


try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
except ImportError:
    print(f"{RED}openpyxl 未安装，请先执行{BLUE} pip3 install openpyxl {RED}。{NC}")
    exit(1)


# 加载Excel文件
try:
    workbook = load_workbook(filePath)
    # 执行其他操作
except InvalidFileException:
    print(f"{RED}无效的文件{NC}")
    exit(1)
except Exception as e:
    print(f"{RED}执行失败:{str(e)} {NC}")
    exit(1)
sheet = workbook.active

# 遍历每一行（从第2行开始）
for row in range(startRowNo, sheet.max_row + 1):
    current_tti = sheet.cell(row=row, column=valueColumnNo).value
    previous_tti = sheet.cell(row=row - 1, column=valueColumnNo).value
    if current_tti is None or previous_tti is None:
        print(f"{RED}您总共要处理{sheet.max_row}行，但第{row}行TTI值不能为空，请检查！{NC}")
        exit(1)
    current_page_class = sheet.cell(row=row, column=idColumnNo).value
    previous_page_class = sheet.cell(row=row - 1, column=idColumnNo).value

    # 检查当前行和前一行的页面类是否相同
    if current_page_class == previous_page_class:
        # 计算当前行和前一行的TTI值
        tti_difference = current_tti - previous_tti

        # 将计算结果写入该行的第7列
        sheet.cell(row=row, column=valueDiffColumnNo).value = tti_difference

        # 检查当前行的TTI是否大于30
        if tti_difference > int(failureMS):
            cellColorString = 'FFFFDCDC'    # 255, 221, 220
        elif tti_difference < int(successMS):
            cellColorString = 'FFD9F2E3'    # 217、242、227
        else:
            cellColorString = 'FFDDDDE3'    # 221、223、227
        fill = PatternFill(start_color=cellColorString, end_color=cellColorString, fill_type='solid')
        for cell in sheet[row]:
            cell.fill = fill
    else: 
        cellColorString = 'FFFFFFFF'
        fill = PatternFill(start_color=cellColorString, end_color=cellColorString, fill_type='solid')
        for cell in sheet[row]:
            cell.fill = fill
    

# 保存修改后的Excel文件
workbook.save(resultSaveToFilePath)


import subprocess
import sys
import platform
def open_file_with_default_tool(file_path):
    if sys.platform.startswith('darwin'):
        subprocess.call(('open', file_path))  # macOS
    elif sys.platform.startswith('win32'):
        subprocess.call(('start', file_path), shell=True)  # Windows
    elif sys.platform.startswith('linux') or platform.system() == 'Linux':
        subprocess.call(('xdg-open', file_path))  # Linux
    else:
        print("Unsupported platform: " + sys.platform)


open_file_with_default_tool(resultSaveToFilePath)